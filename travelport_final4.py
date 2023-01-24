
from logging.config import dictConfig
import requests
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from xml.etree import ElementTree
import xmltodict
import json 
import pandas as pd
from datetime import datetime
import mysql.connector
import re
import numpy as np
from openpyxl import load_workbook
import random

def unique(list1):
    unique_list = []
    for x in list1:
        if x not in unique_list:
            unique_list.append(x)
    return unique_list

def getList(dict):
    list = []
    for key in dict.keys():
        list.append(key)  
    return list

def Convert(string):
    li = list(string.split(","))
    return li

def listToString(s):
   
    # initialize an empty string
    str1 = " "
   
    # return string 
    return (str1.join(s))

new_travelport = pd.read_csv('data.txt', header=None)
new_travelport.columns=['travelport_id']

mydb = mysql.connector.connect(
host="bizly-dev-do-user-12143907-0.b.db.ondigitalocean.com",
database='bizly_dev_20220801',
user="doadmin",
password="AVNS_JeUb_97kvCdsOz8uun-",
port=25060  
)

active_property=pd.read_fwf('ActiveProperties-1G-20221219.txt',colspecs=[(0,4),(4,12),(12,73),(73,174),(174,205),(205,211),(211,215),(215,217),(217,229),(229,-1)],
names=['CHAIN_CD','PPTY_ID','PPTY_NAME','ADDR_LINE','CTY_NAME','CTY_CD','STATE_CD','CNTRY_CD','POSTAL_CD','PHONE_NUM'])

tor_data=pd.read_excel('tor_data.xlsx')

active_property=pd.merge(active_property,tor_data, how='left' ,left_on=['CHAIN_CD','PPTY_ID'], right_on = ['CHAIN_CD','PPTY_ID'])
print(active_property.shape)



#active_property=active_property[active_property['tor_id']==1]
print(active_property.shape)

mycursor=mydb.cursor()
mycursor.execute("SELECT id, state_code FROM bizly_dev_20220801.cities where state_code <>'' ")
myresult = mycursor.fetchall()
data=pd.DataFrame(myresult,columns=['city_id','city_code'])
cities=data

active_property=pd.merge(active_property,cities, how='left' ,left_on=['CTY_CD'], right_on = ['city_code'])

mycursor=mydb.cursor()
mycursor.execute("SELECT id, code  FROM bizly_dev_20220801.countries ")
myresult = mycursor.fetchall()
data=pd.DataFrame(myresult,columns=['country_id','country_code'])
countries=data

active_property=pd.merge(active_property,countries, how='left' ,left_on=['CNTRY_CD'], right_on = ['country_code'])


mycursor=mydb.cursor()  
mycursor.execute("select distinct chain_code,chain_code_id from travelport_venues")
myresult = mycursor.fetchall()
data=pd.DataFrame(myresult,columns=['chain_code','chain_code_id'])
chains=data

active_property=pd.merge(active_property,chains, how='left' ,left_on=['CHAIN_CD'], right_on = ['chain_code'])
print(active_property.shape)


active_property=active_property.astype({'CNTRY_CD':'string'})
active_property=active_property.astype({'STATE_CD':'string'})
active_property=active_property.astype({'POSTAL_CD':'string'}) 
active_property=active_property.astype({'CTY_CD':'string'})
active_property=active_property.astype({'CHAIN_CD':'string'})
active_property=active_property.astype({'PPTY_ID':'string'})  
active_property=active_property.astype({'ADDR_LINE':'string'})
active_property=active_property.astype({'CTY_NAME':'string'})
active_property_all=active_property

active_property=active_property.drop_duplicates(subset=['PPTY_ID'])

print(active_property.shape)

#49500
for prop in range(110000,120000,1): 
#for brojac in range(50):
    #print(brojac)
    #prop=random.randint(0,135000)
    try:
        
        print(active_property.iloc[prop].loc['PPTY_ID'])
        print(active_property.iloc[prop].loc['CHAIN_CD'])

        print(prop)
        start=datetime.now()

        
        text_for_image=""
        without_image=0
        df=[]
        df = pd.DataFrame(columns=['BrandCode', 'BrandName', 'ChainCode','RoommasterPropertyID', 'ChainName', 'CurrencyCode', 'HotelCityCode', 'HotelCode', 'HotelCodeContext', 
        'HotelName', 'LanguageCode','HotelStatus','HotelStatusCode','WhenBuilt','Segment_category','CodeDetail','Quantity','number_meeting_rooms','LargestRoomSpace','LargestSeatingCapacity','MeetingRoomCount','TotalRoomSpace'
        ,'DescriptiveText','Latitude','Longitude','images','caption','CityName','PostalCode','Country','Email','Phones','max_capacity','display_address','full_address'
        ,'primary_contact','images_new','chain_id','city_id'])
            
        form = '''
        <soap:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/">
        <soap:Header/>
        <soap:Body>
        <OTA_HotelDescriptiveInfoRQ TimeStamp="2011-09-07T09:30:47.0Z" Version="1.0" Target="Intrnal" EchoToken="Intenal" xmlns="http://www.opentravel.org/OTA/2003/05">
                <POS>
                <Source>
                    <BookingChannel Type="2">
                        <CompanyName Code="HCP_BIZLY18" CompanyShortName="HCP_BIZLY18"/>     
                    </BookingChannel>
                </Source>
                </POS>
                <HotelDescriptiveInfos>
                    <HotelDescriptiveInfo BrandCode="{bc}" HotelCode="{hc}" HotelCodeContext="1G.EN"/>
                </HotelDescriptiveInfos>
            </OTA_HotelDescriptiveInfoRQ>
        </soap:Body>
        </soap:Envelope> '''.format(bc=active_property.iloc[prop].loc['CHAIN_CD'],hc=active_property.iloc[prop].loc['PPTY_ID'])
        #.format(bc='AK',hc='F9375')
        #Chain_id=active_property.iloc[prop].loc['chain_code_id']
        Chain_id=None
        Property_id=active_property.iloc[prop].loc['PPTY_ID']
        city_id=active_property.iloc[prop].loc['city_id']
        Country_id=active_property.iloc[prop].loc['country_id']

        display_address=active_property.iloc[prop].loc['ADDR_LINE'] + ", " + active_property.iloc[prop].loc['CTY_NAME']
        full_address=active_property.iloc[prop].loc['ADDR_LINE'] + ", " + active_property.iloc[prop].loc['CTY_NAME'] + ", "+active_property.iloc[prop].loc['STATE_CD'] + " " + active_property.iloc[prop].loc['POSTAL_CD'] + ", " + active_property.iloc[prop].loc['CNTRY_CD']
        
        try:
            full_address=full_address.strip().title()
            full_address=full_address.replace("'","%*+")

            display_address=display_address.strip().title()
            display_address=display_address.replace("'","%*+")
        except:
            None

        
        phone2=active_property.iloc[prop].loc['PHONE_NUM']

        try:
            encoded_request = form.encode('utf-8')
            headers = {"Authorization":"Basic SENQL0hDUF9CSVpMWTE4Okh0bEBwbHVzMTg=","Content-Type": "text/xml; charset=UTF-8","Content-Length": "120000000"}
            response = requests.post(url="https://americas.webservices.travelport.com/B2BGateway/connect/HotelContent/HotelDescriptiveInfo",headers = headers,data = encoded_request,verify=False)
            dict_data = xmltodict.parse(response.content)
        except:
            None
        
        try:
            images_all=dict_data['SOAP-ENV:Envelope']['SOAP-ENV:Body']['OTA_HotelDescriptiveInfoRS']['HotelDescriptiveContents']['HotelDescriptiveContent']['MultimediaDescriptions']['MultimediaDescription']['ImageItems']['ImageItem']    
            captions=[]
            categories=[]
            urls=[]
            guest_rooms=[]
            guest_room_num=0
            categories_rooms=['1','2','3','4','5','6','7','8','9','13','23']
            if isinstance(images_all,dict)==True and images_all['@Category'] in ['1','2','3','4','5','6','7','8','9','13','23']:
                caption=images_all['Description']['@Caption']
                caption=caption.replace("'","%*+")
                category=images_all['@Category']
                urls=images_all[0]['@CodeDetail']
                if images_all[i]['@Category'] in ['6','7']:
                    gr=True
                else:
                    gr=False
            else:
                max_width=0
                max_index=0
                i=0
                images_all = sorted(images_all, key=lambda d: d['@Category'])
                for cat_rooms in categories_rooms:
                    cat_rooms_num=0
                    image_cat = [d for d in images_all if d['@Category'] == cat_rooms]           
                    if len(image_cat)>0:
                        df_images=pd.DataFrame(index=range(len(image_cat)),columns=range(9))
                        df_images.columns=['Category', 'num', 'height','width', 'url_last7','XXL_format','caption','url','guestroom']
                        for i in range(len(image_cat)):

                            try:
                                c=image_cat[i]['Description']['@Caption']
                                c=c.replace("'","%*+")
                                c=c.replace(",","/")
                            except:
                                
                                conditions = [
                                (image_cat[i]['@Category']=='1'),
                                (image_cat[i]['@Category']=='2'),
                                (image_cat[i]['@Category']=='3'),
                                (image_cat[i]['@Category']=='4'),
                                (image_cat[i]['@Category']=='5'),
                                (image_cat[i]['@Category']=='6'),
                                (image_cat[i]['@Category']=='7'),
                                (image_cat[i]['@Category']=='8'),
                                (image_cat[i]['@Category']=='9'),
                                (image_cat[i]['@Category']=='13'),
                                (image_cat[i]['@Category']=='23')
                                ]

                                values = ['Exterior', 'Interior', 'Pool view"', 'Restaurant', 'Health club','Guest room','Suite','Meeting room','Ballroom','Bar/Lounge','Bussines center']

                                c = np.select(conditions, values)
                                

                            cat=image_cat[i]['@Category']
                            height=float(image_cat[i]['ImageFormat']['@Height'])
                            width=float(image_cat[i]['ImageFormat']['@Width'])
                            url=image_cat[i]['ImageFormat']['URL']
                            url=url.replace('http://','https://')
                            url_last7=url[-7:]

                            df_images.iloc[i].loc['Category']=cat
                            df_images.iloc[i].loc['num']=i
                            df_images.iloc[i].loc['height']=height
                            df_images.iloc[i].loc['width']=width
                            df_images.iloc[i].loc['url_last7']=url_last7
                            if df_images.iloc[i].loc['url_last7']=='XXL.jpg':
                                df_images.iloc[i].loc['XXL_format']=1
                            else:
                                df_images.iloc[i].loc['XXL_format']=0
                            df_images.iloc[i].loc['caption']=c
                            df_images.iloc[i].loc['url']=url
                            if df_images.iloc[i].loc['Category'] in ['6','7']:
                                df_images.iloc[i].loc['guestroom']=True
                            else:
                                df_images.iloc[i].loc['guestroom']=False

                        df_images=df_images.sort_values(by=['XXL_format','height','width'], ascending=False)

                        df_images=df_images.drop_duplicates(subset = "url")

                        #df_images.to_excel('images.xlsx')

                        max_size= min (3, df_images.shape[0])

                        for i in range(0,max_size,1):
                            captions.append(df_images.iloc[i].loc['caption'])
                            categories.append(df_images.iloc[i].loc['Category'])
                            urls.append(df_images.iloc[i].loc['url'])
                            guest_rooms.append(df_images.iloc[i].loc['guestroom'])

                guest_room_num=any(x == True for x in guest_rooms)
                        
                df_=pd.DataFrame()

                if guest_room_num==True:
                    guest_room_num_="yes"
                else:
                    guest_room_num_="no"

                df_['urls']=urls
                df_['captions']=captions
                df_['guest_rooms']=guest_rooms
                df_['categories']=categories

                
                df_=df_.sort_values(by=['categories'])
                #df_.to_excel('bbb.xlsx')

                urls=df_['urls']
                captions=df_['captions']
                categories=df_['categories']
                guest_rooms=df_['guest_rooms']

                #captions=unique(captions)
                caption = ','.join([str(elem) for elem in captions])
                url = ','.join([str(elem) for elem in urls])
                guest_room=','.join([str(elem) for elem in guest_rooms])
                
        except:
            url=""
            text_for_image="Without image"
        
        if url!="":
            try:
                hotel_content=dict_data['SOAP-ENV:Envelope']['SOAP-ENV:Body']['OTA_HotelDescriptiveInfoRS']['HotelDescriptiveContents']['HotelDescriptiveContent']
                hotel_content_keys=getList(hotel_content)
            except:
                hotel_content_keys=[]

            list_attributes=['@BrandCode', '@BrandName', '@ChainCode', '@ChainName', '@CurrencyCode', '@DecimalPlaces', '@HotelCityCode', '@HotelCode', '@HotelCodeContext', '@HotelName', '@LanguageCode']
            list_fields=list=['BrandCode', 'BrandName', 'ChainCode', 'ChainName', 'CurrencyCode', 'DecimalPlaces', 'HotelCityCode', 'HotelCode', 'HotelCodeContext', 'HotelName', 'LanguageCode']

            for attr,var in zip(list_attributes,list_fields):
                if attr in hotel_content_keys:
                    globals()[var]=hotel_content[attr]
                else:
                    globals()[var]=""

        
            try:
                hotel_info=dict_data['SOAP-ENV:Envelope']['SOAP-ENV:Body']['OTA_HotelDescriptiveInfoRS']['HotelDescriptiveContents']['HotelDescriptiveContent']['HotelInfo']
            except:
                None

            list_attributes=['@HotelStatus', '@HotelStatusCode', '@WhenBuilt']
            list_fields=['HotelStatus','HotelStatusCode','WhenBuilt']

            hotel_info_keys=getList(hotel_info)

            for attr,var in zip(list_attributes,list_fields):
                if attr in hotel_info_keys:
                    globals()[var]=hotel_info[attr]
                else:
                    globals()[var]=""

            try:
                hotel_category=dict_data['SOAP-ENV:Envelope']['SOAP-ENV:Body']['OTA_HotelDescriptiveInfoRS']['HotelDescriptiveContents']['HotelDescriptiveContent']['HotelInfo']['CategoryCodes']['SegmentCategory']
            except:
                None

            try:
                hotel_category=dict_data['SOAP-ENV:Envelope']['SOAP-ENV:Body']['OTA_HotelDescriptiveInfoRS']['HotelDescriptiveContents']['HotelDescriptiveContent']['HotelInfo']['CategoryCodes']['SegmentCategory']
                hotel_categories=[]
                if isinstance(hotel_category,dict)==True:
                    hotel_categories=hotel_category['@CodeDetail']
                else:
                    for i in range(len(hotel_category)):
                        hotel_categories.append(hotel_category[i]['@CodeDetail'])

                    hotel_categories=unique(hotel_categories)
                    hotel_categories = ','.join([str(elem) for elem in hotel_categories])
            except:
                hotel_categories=""

            try:
                GuestRoomInfo=dict_data['SOAP-ENV:Envelope']['SOAP-ENV:Body']['OTA_HotelDescriptiveInfoRS']['HotelDescriptiveContents']['HotelDescriptiveContent']['HotelInfo']['CategoryCodes']['GuestRoomInfo']
                guest_len=len(GuestRoomInfo)
            except:
                GuestRoomInfo=[]
                guest_len=0

            list_attributes=['@CodeDetail']
            list_fields=['CodeDetail']

            CodeDetail=[]

            try:
                GuestRoomInfo=dict_data['SOAP-ENV:Envelope']['SOAP-ENV:Body']['OTA_HotelDescriptiveInfoRS']['HotelDescriptiveContents']['HotelDescriptiveContent']['HotelInfo']['CategoryCodes']['GuestRoomInfo']
                
            except:
                CodeDetail=[]

            try:
                facility_info=dict_data['SOAP-ENV:Envelope']['SOAP-ENV:Body']['OTA_HotelDescriptiveInfoRS']['HotelDescriptiveContents']['HotelDescriptiveContent']['FacilityInfo']['MeetingRooms']
                facility_info_keys=getList(facility_info)

            except:
                facility_info=""
                facility_info_keys=""


            try:
                meeting_rooms=dict_data['SOAP-ENV:Envelope']['SOAP-ENV:Body']['OTA_HotelDescriptiveInfoRS']['HotelDescriptiveContents']['HotelDescriptiveContent']['FacilityInfo']['MeetingRooms']['MeetingRoom']
                number_meeting_rooms=len(meeting_rooms)
            except:
                number_meeting_rooms=""


            list_attributes=['@LargestRoomSpace', '@LargestSeatingCapacity', '@MeetingRoomCount','@TotalRoomSpace']
            list_fields=['LargestRoomSpace','LargestSeatingCapacity','MeetingRoomCount','TotalRoomSpace']

            for attr,var in zip(list_attributes,list_fields):
                if attr in facility_info_keys:
                    globals()[var]=facility_info[attr]
                else:
                    globals()[var]=""

            try:
                DescriptiveText=dict_data['SOAP-ENV:Envelope']['SOAP-ENV:Body']['OTA_HotelDescriptiveInfoRS']['HotelDescriptiveContents']['HotelDescriptiveContent']['HotelInfo']['Descriptions']['DescriptiveText']
                DescriptiveText=DescriptiveText.replace("'","%*+")
                DescriptiveText=DescriptiveText.replace("\xa0"," ")
                DescriptiveText=DescriptiveText.replace("\u00a0","")
                DescriptiveText=DescriptiveText.replace('"','')
            except:
                DescriptiveText=""


            try:
                position=dict_data['SOAP-ENV:Envelope']['SOAP-ENV:Body']['OTA_HotelDescriptiveInfoRS']['HotelDescriptiveContents']['HotelDescriptiveContent']['HotelInfo']['Position']
            except:
                None

            list_attributes=['@Latitude','@Longitude']
            list_fields=['Latitude','Longitude']

            for attr,var in zip(list_attributes,list_fields):
                if attr in position:
                    globals()[var]=position[attr]
                else:
                    globals()[var]=""

            
            
            try:
                address=dict_data['SOAP-ENV:Envelope']['SOAP-ENV:Body']['OTA_HotelDescriptiveInfoRS']['HotelDescriptiveContents']['HotelDescriptiveContent']['ContactInfos']['ContactInfo']['Addresses']['Address']
            except:
                address=""

            list_attributes=['CityName','PostalCode','Code']
            list_fields=['CityName','PostalCode','Country']

            for attr,var in zip(list_attributes,list_fields):
                if attr in list_attributes:
                    try:
                        globals()[var]=address[attr]
                    except:
                        globals()[var]=""
                else:
                    globals()[var]=""

                try:
                    if Country=="":
                        Country=address['CountryName']['@Code']
                except:
                    Country=""

            try:
                emails=dict_data['SOAP-ENV:Envelope']['SOAP-ENV:Body']['OTA_HotelDescriptiveInfoRS']['HotelDescriptiveContents']['HotelDescriptiveContent']['ContactInfos']['ContactInfo']['Emails']['Email']
                email=emails['#text']
            except:
                email=""

            if email=="" :
                try:
                    email=dict_data['SOAP-ENV:Envelope']['SOAP-ENV:Body']['OTA_HotelDescriptiveInfoRS']['HotelDescriptiveContents']['HotelDescriptiveContent']['HotelInfo']['Descriptions']['MultimediaDescriptions']['MultimediaDescription']['TextItems']['TextItem'][0]['Description']['#text']
                    email=str(email)
                    email = re.search(r'[\w.+-]+@[\w-]+\.[\w.-]+', email)
                    if email.endswith(('.com','.org','.gov','.edu','.net','.mil')):
                        email=email.group(0)
                    else:
                        email=""
                except:
                    email=""
            
            phone=""
            try:
                phones=dict_data['SOAP-ENV:Envelope']['SOAP-ENV:Body']['OTA_HotelDescriptiveInfoRS']['HotelDescriptiveContents']['HotelDescriptiveContent']['ContactInfos']['ContactInfo']['Phones']['Phone']
            except:
                phone=""

            try: 
                phone_number=phones[0]['@PhoneNumber']
                area_city_code=phones[0]['@AreaCityCode']
                country_access_code=phones[0]['@CountryAccessCode']
                phone=country_access_code+'-'+area_city_code+'-'+phone_number
            except:
                None

            if phone=="":
                try:
                    phone=phones[0]['@PhoneNumber']
                except:
                    phone=""

            if phone=="":
                phone=phone2


            try:
                Quantity=dict_data['SOAP-ENV:Envelope']['SOAP-ENV:Body']['OTA_HotelDescriptiveInfoRS']['HotelDescriptiveContents']['HotelDescriptiveContent']['FacilityInfo']['MeetingRooms']['@MeetingRoomCount']
                max_capacity=str(Quantity)
            except:
                Quantity=0
                max_capacity="unknown"
            
            data2=pd.DataFrame(index=range(1),columns=range(5))
            data2.columns=['email','first_name','last_name','phone','travelport_id']

            data2['email']=email
            data2['first_name']=None
            data2['last_name']=None
            data2['phone']=phone
            data2['travelport_id']=Property_id

            url_new=url.split(",")
            caption_new=caption.split(",")
            guest_room_new=guest_room.split(",")
            
            data3=pd.DataFrame(index=range(len(url_new)),columns=range(5))
            data3.columns=['name','description','src_url','guestroom','travelport_id']
            for im in range(len(url_new)):

                try:
                    if len(url_new)==len(caption_new):
                        data3.iloc[im].loc['name']=caption_new[im]
                    else:
                        data3.iloc[im].loc['name']=""
                except:
                    data3.iloc[im].loc['name']=""
                data3.iloc[im].loc['description']=None
                data3.iloc[im].loc['src_url']=url_new[im]
                data3.iloc[im].loc['guestroom']=guest_room_new[im]
                if data3.iloc[im].loc['guestroom']=='False':
                    data3.iloc[im].loc['guestroom']=False
                else:
                    data3.iloc[im].loc['guestroom']=True

                data3.iloc[im].loc['travelport_id']=Property_id
            
            
            primary_contact = (data2.groupby(['travelport_id'])
                    .apply(lambda x: x[['email','first_name','last_name','phone']].to_dict('records'))
                    .reset_index()
                    .rename(columns={0: 'primary_contact'}))


            images_new=(data3.groupby(['travelport_id'])
                .apply(lambda x: x[['name','description','src_url','guestroom']].to_dict('records'))
                .reset_index()
                .rename(columns={0: 'images_new'}))

            df.loc[len(df.index)] = [BrandCode, BrandName, ChainCode,Property_id, ChainName, CurrencyCode, HotelCityCode, HotelCode, HotelCodeContext, 
            HotelName, LanguageCode,HotelStatus,HotelStatusCode,WhenBuilt,hotel_categories,CodeDetail,Quantity,number_meeting_rooms,LargestRoomSpace,LargestSeatingCapacity,
            MeetingRoomCount,TotalRoomSpace,DescriptiveText,Latitude,Longitude,url,caption,CityName,PostalCode,Country_id,email,phone,max_capacity,display_address,full_address,
            primary_contact['primary_contact'].values[0],images_new['images_new'].values[0],Chain_id,city_id]
        else:
            text_for_image="Without image"
            
        
        #df.to_excel('aaa.xlsx')

        data = pd.DataFrame(index=range(df.shape[0]),columns=range(20))
        data.columns=['name','chain_id','summary','type_id','catering','av','room_count','max_capacity','external_url','display_address',
        'full_address','google_place_id','city_id','lat','lng','travelport_id','source','country_id','primary_contact','images']

        data['name']=df['HotelName'].astype('string').str.strip().str.title()
        data['chain_id']=df['chain_id']
        data['summary']=df['DescriptiveText']
        data['type_id']=1
        data['catering']=None
        data['av']=True
        data['room_count']=df['max_capacity']
        data['max_capacity']=df['LargestRoomSpace']
        data['external_url']=None
        data['display_address']=df['display_address']
        data['full_address']=df['full_address']
        data['google_place_id']=None
        data['city_id']=df['city_id']
        data['lat']=df['Latitude']
        data['lng']=df['Longitude']
        data['travelport_id']=df['RoommasterPropertyID']
        data['source']='travelport'
        data['country_id']=Country_id
        data['primary_contact']=df['primary_contact']
        data['images']=df['images_new']

        data=data[['name','chain_id','summary','type_id','catering','av','room_count','max_capacity','external_url','display_address',
        'full_address','google_place_id','city_id','lat','lng','travelport_id','source','country_id','primary_contact','images']] 
        #data.to_excel('aaa.xlsx')

        n_end_sample=data.shape[0]
        try:
            if int(max_capacity)>0:
                with_meeting_room='yes'
            else:
                with_meeting_room='no'
        except:
            with_meeting_room="unknown"

        data.to_json("test.json", orient="records")
        data_=data.to_json("test.json", orient="records")
        data_ = json.load(open("test.json"))


        final_str=''
        l=0
        for l in range(len(data_)):
            if l==0:
                final_str=str(data_[l])
            else:
                final_str=final_str+','+ str(data_[l])
        final_data='{ "venues":['+final_str+']}'
        #final_data=final_data.replace("'","%*+")
        final_data=final_data.replace("'",'"')
        final_data=final_data.replace("None","null")

        final_data=final_data.replace("True","true")
        final_data=final_data.replace("False","false")

        final_data=final_data.replace("%*+","'")
        #final_data=final_data.replace("-","")
        #try:
        json_object = json.loads(final_data)
        with open('sample.json', 'w') as f:
            json.dump(json_object, f)
        #check new data type

        

        #final_data = final_data.encode('utf-8')
        headers = {"Content-Type": "application/json","Content-Length": "120000000","bizly-api-client-id": "ci-prod_milos-travelport", "bizly-api-client-secret": "sk-prod_MTP-sna81bn8sv843as8asi2580asom112"}
        #headers = {"bizly-api-client-id": "ci-prod_milos-travelport", "bizly-api-client-secret": "sk-prod_MTP-sna81bn8sv843as8asi2580asom112"}
        response = requests.post(url="https://api.bizly.com/venues/bulk-upload", headers=headers,json =json_object)
        #response = requests.post(url="https://api.bizly.com/venues/bulk-upload", headers=headers,data =form)
        #response = requests.get(url="https://api.bizly.com/venues/bulk-upload")
        if response.status_code==200:
            msg=''
        else:
            msg=response.content
        print(msg)

        end=datetime.now()

        minutes=(end-start)
        minutes=minutes.seconds / 60


        wb = load_workbook('Travelport4.xlsx')
        ws = wb['Statistics']

        #ws.cell(row,1).value = minutes



        for cell in ws["A"]:
            if cell.value is None:
                row= cell.row
                break
        else:
            row= cell.row+1

        ws.cell(row,1).value=str(active_property.iloc[prop].loc['PPTY_ID'])
        ws.cell(row,2).value=str(active_property.iloc[prop].loc['CHAIN_CD'])
        ws.cell(row,3).value=text_for_image
        ws.cell(row,4).value=with_meeting_room
        ws.cell(row,5).value=guest_room_num_
        ws.cell(row,6).value=minutes
        ws.cell(row,7).value=response.status_code
        ws.cell(row,8).value=msg
        ws.cell(row,9).value=final_data

        wb.save('Travelport4.xlsx')

        import os
        os.remove('sample.json')

    except:
        wb = load_workbook('Travelport4.xlsx')
        ws = wb['Statistics']
        for cell in ws["A"]:
            if cell.value is None:
                row= cell.row
                break
        else:
            row= cell.row+1

        ws.cell(row,1).value=str(active_property.iloc[prop].loc['PPTY_ID'])
        ws.cell(row,2).value=str(active_property.iloc[prop].loc['CHAIN_CD'])

        
        wb.save('Travelport4.xlsx')

